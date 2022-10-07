#from add_capacity_CA import add_capacity
from collections import defaultdict
from openpyxl.workbook import workbook
import pandas as pd
import shutil
import math
import time
from openpyxl import load_workbook
#import xlwings as xw
import os
#pip install geopy in requirements
from geopy.distance import geodesic

#### Global Variables #####
periods = ["2025", "2030", "2035", "2040", "2045", "2050"]
provinces_full = {
            'BC': "British Columbia",
            'AB': "Alberta",
            'MN': "Manitoba",
            'NB': "New Brunswick",
            'NL': "Newfoundland and Labrador",
            'NS': "Nova Scotia",
            'ON': "Ontario",
            'QB': "Quebec",
            'SK': "Saskatchewan",
            'PE': "Prince Edward Island"
            }
### Add the provinces that you'd like to be extracted from the COPPER output
provinces=list(provinces_full.keys())
provinces=['AB']

"""
    This function ensures that any forumulas from the Excel sheet are read
    as values by Pandas and not as an empty cell.
"""
def df_from_excel(path, sheet_name):
    # app = xw.App(visible=False)
    # book = app.books.open(path)
    # book.save()
    # app.kill()
    
    return pd.read_excel(path, header=0, sheet_name=sheet_name)

"""
    This function calculates the distance between two points.
"""
def get_distance(point_A, point_B):
    lat1 = point_A[0]
    lon1 = point_A[1]
    lat2 = point_B[0]
    lon2 = point_B[1]

    dist = math.sqrt((lat2 - lat1)**2 + (lon2 - lon1)**2)
    return dist

"""
    This function is taken from add_capacity_CA and modified
"""
def add_vre_capacity():
    print("##### Adding vre capacity ######")
    for province in provinces:
        print(f'Calculating VRE for {province}')
        for i,period in enumerate(periods):
        
            if i == 0:
                shutil.copyfile(f"model inputs - {province}.xlsx", f"model inputs - {province}_{period}.xlsx")
            else:
                shutil.copyfile(f"model inputs - {province}_{periods[i-1]}.xlsx", f"model inputs - {province}_{period}.xlsx")
                
            mock = df_from_excel(f"model inputs - {province}_{period}.xlsx", "vre plants")
            vre_model_input_df = pd.DataFrame([], columns =mock.columns)
            
            wb_target = load_workbook(f"model inputs - {province}_{period}.xlsx", data_only=True)
            del wb_target["vre plants"]
            writer = pd.ExcelWriter(f"model inputs - {province}_{period}.xlsx", engine='openpyxl')
            writer.book = wb_target
            vre_model_input_df.to_excel(writer, sheet_name="vre plants", index=False)
            writer.save()
            wb_target.close()

        for gen_type in ["solar", "wind"]:
            capacity_df = pd.read_csv(f"capacity_{gen_type}.csv", header=None)
            capacity_recon_df = pd.read_csv(f"capacity_{gen_type}_recon.csv", header=None)
            extant_df = pd.read_csv(f"extant_{gen_type}.csv", header=None)
            coordinate_df = pd.read_excel("coordinate.xlsx", sheet_name="coordinate_system")
            gl_to_ba_df = pd.read_csv("map_gl_to_ba.csv", header=None)
            bus_df = pd.read_excel(f"210818-{province}-DataInventory.xlsx", sheet_name="Nodes", header=1)
            
            capacity_count_df = pd.read_csv(f"capacity_{gen_type}.csv", header=None)
            
            prev_period = periods[0]
            for index,row in capacity_df.iterrows():
                capacity_df.at[index, 2] = capacity_df.at[index, 2] + capacity_recon_df.at[index, 2]
                capacity_count_df.at[index, 2] = capacity_df.at[index, 2] + capacity_recon_df.at[index, 2] + extant_df.at[index, 2]
                
            for index,row in capacity_df.loc[capacity_count_df[2] != 0].iterrows():
                
                period_f = row[0].replace('(','').replace("'", '')
                
                if prev_period != period_f and gen_type == "solar":
                    period_index = periods.index(prev_period)
                    vre_model_input_df = df_from_excel(f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                elif prev_period != period_f and gen_type == "wind":
                    period_index = periods.index(period_f)
                    vre_model_input_df_solar = df_from_excel(f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                    period_index = periods.index(prev_period)
                    vre_model_input_df = df_from_excel(f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                    
                    vre_model_input_df = pd.concat([vre_model_input_df_solar, vre_model_input_df], ignore_index=True)
                else:
                    period_index = periods.index(period_f)
                    vre_model_input_df = df_from_excel(f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                
                vre_model_input_df = vre_model_input_df.drop_duplicates(subset=['plant ID'])
                grid_cell = row[1].replace(')','').replace("'", '')

                if(gl_to_ba_df.at[index%2278, 1].split(".")[0] == provinces_full[province]):
                    
                    capacity = extant_df.at[index, 2] + capacity_df.iat[index, 2]
                    
                    latitude = None
                    longitude = None
                    balancing_area = None
                    
                    for j,row in coordinate_df.loc[coordinate_df['grid cell'] == int(grid_cell)].iterrows():
                        latitude = row['lat']
                        longitude = row['lon']
                
                    for j,row in gl_to_ba_df.loc[gl_to_ba_df[0] == int(grid_cell)].iterrows():
                        balancing_area = row[1]
                    
                    coords_1 = (latitude, longitude)
                    coords_2 = (bus_df.at[2, "Latitude"], bus_df.at[2, "Longitude"])
                    min = geodesic(coords_1, coords_2).km
                    i = 0
                    for j,row in bus_df.iterrows():
                        if(str(row["Latitude"]) != "nan"):
                            coords_2 = (row["Latitude"], row["Longitude"])
                            temp = geodesic(coords_1, coords_2).km
                            if(min > temp):
                                min = temp
                                i = j
                    
                    bus = bus_df.at[i, "Node Name"]
                    
                    temp = vre_model_input_df.loc[vre_model_input_df["latitude"] == latitude]
                    temp = temp.loc[temp['kind'] == gen_type]
                    
                    if len(temp.loc[temp['longitude'] == longitude]) == 0:
                        
                        suffix = 1
                        name = f"{gen_type[0].upper()}{gen_type[1:]}_{suffix}" 
                        
                        while len(vre_model_input_df.loc[vre_model_input_df['name'] == name]) != 0:
                            suffix += 1
                            name = f"{gen_type[0].upper()}{gen_type[1:]}_{suffix}" 
                        merra_lat = round(2*(latitude+90),0)
                        merra_long = round(1.5*(longitude+180),0)                        
                        values = [name, name, gen_type, latitude, longitude, capacity, bus, None, None, merra_lat, merra_long]
                        
                        new_row = pd.DataFrame([values], columns=vre_model_input_df.columns)
                        vre_model_input_df = pd.concat([vre_model_input_df, new_row], ignore_index=True)
                        del temp
                    else:

                        # temp = vre_model_input_df.loc[vre_model_input_df["latitude"] == latitude]
                        temp = temp.loc[temp['longitude'] == longitude]

                        # for j,row in temp.iterrows():
                            # temp = temp.loc[temp['longitude'] == longitude]
                            
                        for g,row in temp.iterrows():
                            
                            vre_model_input_df.at[g, "[MW]"] += capacity
                            if vre_model_input_df.at[g, "[MW]"] < 0:
                                vre_model_input_df.at[g, "[MW]"] = 0
                        del temp
            
                wb_target = load_workbook(f"model inputs - {province}_{period_f}.xlsx", data_only=True)
                #wb_target = load_workbook(f"model inputs - {province}_{periods[i-1]}.xlsx", data_only=True)
                del wb_target["vre plants"]
                writer = pd.ExcelWriter(f"model inputs - {province}_{period_f}.xlsx", engine='openpyxl')
                #writer = pd.ExcelWriter(f"model inputs - {province}_{periods[i-1]}.xlsx", engine='openpyxl')
                writer.book = wb_target
                vre_model_input_df.to_excel(writer, sheet_name="vre plants", index=False)
                writer.save()
                wb_target.close() 
                del vre_model_input_df
                prev_period = period_f

def add_non_VRE():
    result_summary_df = pd.read_excel("Results_summary.xlsx", sheet_name="ABA_generation_mix")
    province_results = {}
    
    for province in provinces:
        result_columns = []
        i = -1
        for columnName, columnData in result_summary_df.iteritems():
            if columnName.split('.')[0] == provinces_full[province]:
                if columnName.split('.')[1] == "b":
                    result_columns[i] += columnData
                else:
                    result_columns.append(columnData)
                    i += 1
        province_results[province] = result_columns
               
    print("##### Adding to Non-VRE #####")
    for province in provinces:
        print(f'Calculating non-VRE for {province}')
        for i, res in enumerate(province_results[province]):
            period = periods[i]
            model_input_df = df_from_excel(f"model inputs - {province}_{period}.xlsx", "non-vre plants")
            
            for index, k in enumerate(res):
                result = k
                gen_type = result_summary_df.iat[index, 0]
                if gen_type in ("solar", "wind"):
                    continue
                if gen_type == "gasCC":
                    gen_type = "NG_CT"
                if gen_type == "gasSC":
                    gen_type = "NG_CG"
                if gen_type == "gasccs":
                    gen_type = "NG_CC"
                if gen_type == "biomass":
                    gen_type = "biogas"
                if gen_type == "hydro":
                    gen_type = "hydro_daily"
                
                k -= model_input_df.loc[model_input_df['kind'] == gen_type]["[MW]"].sum()
                while k != 0:
                    smallest_capacity = None
                    generator_to_remove = None
                    for j,generator in model_input_df.loc[model_input_df['kind'] == gen_type].loc[model_input_df["[MW]"] > 0].iterrows():
                        if smallest_capacity is None:
                            smallest_capacity = generator.get(5)
                            generator_to_remove = j
                        elif smallest_capacity > generator.get(5):
                            smallest_capacity = generator.get(5)
                            generator_to_remove = j
                        else:
                            continue
                    
                    if smallest_capacity is None:
                        break
                    elif k < 0:
                        if abs(k) > smallest_capacity:
                            model_input_df.at[generator_to_remove, "[MW]"] = 0
                            k = k + smallest_capacity
                        else:
                            model_input_df.at[generator_to_remove, "[MW]"] = smallest_capacity + float(k)
                            k = 0
                    elif k >= 0:
                        model_input_df.at[generator_to_remove, "[MW]"] += float(k)
                        k = 0
                    
            model_input_df = model_input_df.loc[model_input_df["[MW]"] > 0]
                    
            ### Write the updates
            wb_target = load_workbook(f"model inputs - {province}_{period}.xlsx", data_only=True)
            del wb_target["non-vre plants"]
            writer = pd.ExcelWriter(f"model inputs - {province}_{period}.xlsx", engine='openpyxl')
            writer.book = wb_target
            model_input_df.to_excel(writer, sheet_name="non-vre plants", index=False)
            writer.save()

def add_storage(province):
    print("##### Adding storage ######")
    capacity_storage_df = pd.read_csv("capacity_storage.csv", header=None)
    capacity_storage_df = capacity_storage_df.sort_values(by=[0])

    for i,row in capacity_storage_df.loc[capacity_storage_df[3] != 0].iterrows(): 
    
        period = str(row.get(0))
        prev_period = periods.index(period) - 1
        balancing_area = row.get(2)
        province_storage = balancing_area.split('.')[0]
        new_capacity = row.get(3)
        storage_type = row.get(1)
        name = storage_type + "." + balancing_area.split('.')[1]
        if provinces_full[province] != province_storage:
            continue
        model_input_df = df_from_excel(f"model inputs - {province}_{periods[prev_period]}.xlsx", "storage")
        
        smallest_capacity = None
        effected_storage = None

        for j,storage in model_input_df.iterrows():
            if storage_type not in storage.get(0):
                continue
            if smallest_capacity is None:
                smallest_capacity = storage.get(4)
                effected_storage = j
            else:
                if smallest_capacity > storage.get(4):
                    effected_storage = j
                    smallest_capacity = storage.get(4)
        
        if effected_storage is None:
            kind = storage_type
            initial = 1
            storage_type = f"{storage_type}_{balancing_area}"
            while len(model_input_df.loc[model_input_df["plant ID"] == type] != 0):
                initial +=1
                storage_type = f"{storage_type}_{balancing_area}"
            storage_capacity_max = 0
            if(kind == "PHS"):
                storage_capacity_max = new_capacity*8
            elif(kind == "LB"):
                storage_capacity_max = new_capacity*4
                
            values = [name, name, kind, None, new_capacity, storage_capacity_max, 1, None, None, "0P", None, None, None, 0]
            
            new_row = pd.DataFrame([values], columns=model_input_df.columns.tolist())
            model_input_df = pd.concat([model_input_df, new_row], ignore_index=True)
        
            
        else:
            model_input_df.at[effected_storage, "[MW]"] = model_input_df.at[effected_storage, "[MW]"] +new_capacity
            if(kind == "PHS"):
                model_input_df.at[effected_storage, "storagecapacitymax"] = model_input_df.at[effected_storage, "[MW]"]*8
            elif(kind == "LB"):
                model_input_df.at[effected_storage, "storagecapacitymax"] = model_input_df.at[effected_storage, "[MW]"]*4
            
        wb_target = load_workbook(f"model inputs - {province}_{period}.xlsx", data_only=True)
        del wb_target["storage"]
        writer = pd.ExcelWriter(f"model inputs - {province}_{period}.xlsx", engine='openpyxl')
        writer.book = wb_target
        model_input_df.to_excel(writer, sheet_name="storage", index=False)
        writer.save()
        
def main():
    print('You need to have these files already available in the folfer:\n1-Coordinates.xls,\n2-map_gl_to_ba.csv,\n'
          '3-all extant files\n4-Data inventory files from CODERS database\n')
    new_path = input("Enter the directory's name containing input files:\n(i.e. CA_date_use)\n")
    os.chdir(os.getcwd() + "/" +  new_path)
    start_time = time.time()
    
    add_vre_capacity()
    add_non_VRE()
    for province in provinces:
        add_storage(province)
    print("--- %s seconds ---" % (time.time() - start_time))
    
if __name__ == "__main__":
    main()