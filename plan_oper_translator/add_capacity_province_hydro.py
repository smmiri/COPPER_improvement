#from add_capacity_CA import add_capacity
from collections import defaultdict
from platform import java_ver
from tkinter import W
from openpyxl.workbook import workbook
import pandas as pd
import numpy as np
from pathlib import Path
import shutil
import math
import time
from time import process_time
from openpyxl import load_workbook
#import xlwings as xw
import os
#pip install geopy in requirements
from geopy.distance import geodesic

# TODO: Could add function to move needed files into COPPER output folder

#### Global Variables #####

periods = ["2030", "2040", "2050"]

# Which provinces to produce inputs for
provinces=['AB']#, 'BC', 'MB', 'SK']

provinces_full = {
            'BC': "British Columbia",
            'AB': "Alberta",
            'MB': "Manitoba",
            'NB': "New Brunswick",
            'NL': "Newfoundland and Labrador",
            'NS': "Nova Scotia",
            'ON': "Ontario",
            'QC': "Quebec",
            'SK': "Saskatchewan",
            'PE': "Prince Edward Island"
            }

# Paths
files_needed_for_tool = Path().cwd() / 'files_needed_for_tool'
copper_path = files_needed_for_tool / 'copper'
coders_path = files_needed_for_tool / 'coders'
model_inputs_in = files_needed_for_tool / 'model_inputs'
model_inputs_out = Path().cwd() / 'results' / 'model_inputs'
output_path = Path().cwd() / 'copper_outputs' / 'outputs2410_ct170_270_370_rd24_pds3_Hr_NoOBPS_LGP_Hydro_NCL_NoPCL_CPHy_NoAr_SMR_CCS_CPO_GPS_TE'
# output_path = Path().cwd() / 'copper_outputs' / input("Enter the name of the COPPER output folder you would like to convert: ")
hydro_out = Path().cwd() / 'results' / 'hydro'

### Add the provinces that you'd like to be extracted from the COPPER output
# provinces=list(provinces_full.keys())

"""
    This function ensures that any forumulas from the Excel sheet are read
    as values by Pandas and not as an empty cell.
"""
def df_from_excel(path, sheet_name):
    # app = xw.App(visible=False)
    # book = app.books.open(path)
    # book.save()
    # app.kill()
    
    return pd.read_excel(path, header=0, sheet_name=sheet_name)

"""
    This function calculates the distance between two points.
"""
def get_distance(point_A, point_B):
    lat1 = point_A[0]
    lon1 = point_A[1]
    lat2 = point_B[0]
    lon2 = point_B[1]

    dist = math.sqrt((lat2 - lat1)**2 + (lon2 - lon1)**2)
    return dist

"""
    This function filters CODERS non-vre plant data
"""
def plants_filter(data_inventory_df):

    filter_df = data_inventory_df[['Project Name', 'Latitude', 'Longitude', 'Installed Capacity', 'Generation Type - COPPER']].dropna()
    filter_df = filter_df.loc[filter_df['Generation Type - COPPER'].str.contains('hydro')].reset_index(drop=True)

    # Aggregate numbered plants
    filter_df['Project Name'] = filter_df['Project Name'].str.partition('_')[0]
    filter_df = filter_df.groupby(['Project Name', 'Generation Type - COPPER', 'Latitude', 'Longitude'], as_index=False)['Installed Capacity'].sum()

    return filter_df

def month(x):
    if x in range(1, 32):
        return 'January'
    elif x in range(32, 61):
        return 'February'
    elif x in range(61, 93):
        return 'March'
    elif x in range(93, 124):
        return 'April'
    elif x in range(124, 156):
        return 'May'
    elif x in range(156, 187):
        return 'June'
    elif x in range(187, 219):
        return 'July'
    elif x in range(219, 251):
        return 'August'
    elif x in range(251, 282):
        return 'September'
    elif x in range(282, 314):
        return 'October'
    elif x in range(314, 345):
        return 'November'
    elif x in range(345, 366):
        return 'December'

def process_day(x):
    return (month(math.floor((x-1)/24)+1), math.floor((x-1)/24)+1, x)

"""
    This function creates a capacity factor dataframe from COPPER data
"""
def hydro_capacities():

    # Read in capacity factors
    df = pd.read_csv(copper_path / 'hydro_cf.csv', names=['province.hour', 'value'])

    # Split up province and hour columns, pivot into new dataframe with provinces as columns
    df[['province', 'hour']] = df['province.hour'].str.split('.', expand=True).copy()
    hydro_cf = df[['hour', 'province', 'value']].copy()
    hydro_cf = hydro_cf.pivot(index = 'hour', columns = 'province', values = 'value').reset_index()
    hydro_cf.index = hydro_cf['hour'].astype('int64')
    hydro_cf.sort_index(inplace=True)
    hydro_cf.drop('hour', axis=1, inplace=True)

    # Group index into months, days of year
    hydro_cf.index = pd.MultiIndex.from_tuples([process_day(x) for x in hydro_cf.index.values], names = ['month', 'day', 'hour'])

    return hydro_cf
"""
    This function formats CODERS nodes to match names in SILVER
"""
def node_formatter(province):

    # Matching node data from CODERS with format from SILVER

    # data = Path().cwd() / "files_needed_for_tool"    

    nodes_format = pd.read_excel(model_inputs_in / f"model inputs - {province}.xlsx", "demand centres")['bus'].drop_duplicates()
    nodes_coders = pd.read_excel(coders_path / f'210818-{province}-DataInventory.xlsx', sheet_name='Nodes', header = 1).dropna(how='all')[['Node Name', 'Latitude', 'Longitude']]

    if province == 'AB':

        # Reformatting data
        nodes_coders['Node Name'] = nodes_coders['Node Name'].str.upper()
        nodes_coders['Node Name'] = nodes_coders['Node Name'].str.replace(' ', "")
        nodes = pd.merge(nodes_coders, nodes_format, left_on='Node Name', right_on='bus')
        nodes = pd.concat([nodes, nodes_coders.loc[nodes_coders['Node Name'].isin(['CASTLEDOWNS557S','DOME665S', 'GENESEE330P', 'NEWELL2075S', 'PETROLIA816S'])]])

        # Renaming to match demand centres
        nodes.loc[nodes['Node Name'] == 'CASTLEDOWNS557S', 'Node Name'] = 'CASTLEDOWNS'
        nodes.loc[nodes['Node Name'] == 'DOME665S', 'Node Name'] = 'DOME'
        nodes.loc[nodes['Node Name'] == 'GENESEE330P', 'Node Name'] = 'GENESEE330'
        nodes.loc[nodes['Node Name'] == 'NEWELL2075S', 'Node Name'] = 'NEWELL2075S89S'
        nodes.loc[nodes['Node Name'] == 'PETROLIA816S', 'Node Name'] = 'PETROLIA'

        nodes['bus'] = nodes['Node Name']
        nodes = nodes.drop('Node Name', axis=1).reset_index(drop=True)


    elif province == 'BC':
       # Add acronyms to beginnings of CODERS names
        for index, value in nodes_format.iteritems():

            remove_acronym = value.partition(' - ')[2]

            nodes_coders.loc[nodes_coders['Node Name'] == remove_acronym, 'Node Name'] = value

        nodes = pd.merge(nodes_coders, nodes_format, left_on='Node Name', right_on='bus')
        nodes = pd.concat([nodes, nodes_coders.loc[nodes_coders['Node Name'].isin(['Kennedy'])]])

        nodes.loc[nodes['Node Name']=='Kennedy', 'bus'] = 'KDY - Kennedy Capacitor Station'
        nodes = nodes.drop('Node Name', axis=1).reset_index(drop=True).sort_index()

    elif province == 'MB':

        nodes_format = nodes_format.str.strip()
        nodes_coders['Node Name'] = nodes_coders['Node Name'].str.upper()
        nodes_coders['Node Name'] = nodes_coders['Node Name'].str.strip() 
        nodes = pd.merge(nodes_coders, nodes_format, left_on='Node Name', right_on='bus')
        nodes = pd.concat([nodes, nodes_coders.loc[nodes_coders['Node Name'] == 'KELSEY TS']])
        
        nodes.loc[nodes['Node Name'] == 'KELSEY TS', 'Node Name'] = 'KELSEY TERMINAL'
        nodes['bus'] = nodes['Node Name']
        nodes = nodes.drop('Node Name', axis=1).reset_index(drop=True)

    # Saskatchewan nodes are already in a matching format
    elif province == 'SK':

        nodes = pd.merge(nodes_coders, nodes_format, left_on='Node Name', right_on='bus').drop('Node Name', axis = 1).reset_index(drop=True)
        nodes = pd.concat([nodes, nodes_coders.loc[nodes_coders['Node Name'] == 'Regina South']])
        nodes['bus'] = nodes['Node Name']
        nodes = nodes.drop('Node Name', axis=1).reset_index(drop=True)
        pass

    
    if len(nodes) != len(nodes_format):
        print("Some nodes were lost: \n")
        # print(nodes['bus'].compare(nodes_format))
    
    return nodes
"""
    Minimum distance finding function, takes one plant dataframe and a dataframe of all nodes, and loops through the nodes to find the closest one
"""
def min_finder(plant_df, nodes_df):

    R = 6371 # Radius of earth
    # Haversine Formula
    d = lambda lat1, lat2, long1, long2: 2*R*np.arcsin(np.sqrt((np.sin((lat2-lat1)/2))**2 + (np.cos(lat1)*np.cos(lat2)*(np.sin((long2-long1)/2))**2)))

    plant_lat = plant_df['Latitude']*np.pi/180
    plant_long = plant_df['Longitude']*np.pi/180

    node_lat = nodes_df.iloc[0]['Latitude']*np.pi/180
    node_long = nodes_df.iloc[0]['Longitude']*np.pi/180

    min = d(plant_lat, node_lat, plant_long, node_long)
    distances = []

    min_index = 0

    for index, node in nodes_df.iterrows():

        node_lat = node['Latitude']*np.pi/180
        node_long = node['Longitude']*np.pi/180

        distance = d(plant_lat, node_lat, plant_long, node_long)
        distances.append(distance)

        if distance < min:
            min = distance
            min_index = index
    
    return min, min_index

def clear_vre(sheet):

    while(sheet.max_row > 1):

        sheet.delete_rows(2)

    return
"""
    This function is taken from add_capacity_CA and modified
"""
# TODO: This function is slow currently, could be sped up
def add_vre_capacity():
    print("##### Adding vre capacity ######")
    for province in provinces:
        print(f'Calculating VRE for {province}')

        # Geographic data for MERRA lat/long, mapping
        coordinate_df = pd.read_excel(files_needed_for_tool / "coordinate.xlsx", sheet_name="coordinate_system")
        gl_to_ba_df = pd.read_csv(files_needed_for_tool / "map_gl_to_ba.csv", header=None)

        # Copy model inputs file for given province, clear data in vre plants sheet
        for i,period in enumerate(periods):
    
            if i == 0:
                # Initial copy of model inputs file
                shutil.copyfile(model_inputs_in / f"model inputs - {province}.xlsx", model_inputs_out / f"model inputs - {province}_{period}.xlsx")
            else:
                shutil.copyfile(model_inputs_out / f"model inputs - {province}_{periods[i-1]}.xlsx", model_inputs_out / f"model inputs - {province}_{period}.xlsx")
                
            # Read model inputs using openpyxl
            model_inputs = load_workbook(model_inputs_out / f"model inputs - {province}_{period}.xlsx", data_only=True)
            vre_inputs = model_inputs['vre plants']

            # Delete contents of vre plants sheet (leaving column headers)
            vre_inputs.delete_rows(2, vre_inputs.max_row - 1)

            # Overwrite original file
            model_inputs.save(model_inputs_out / f"model inputs - {province}_{period}.xlsx")

        # NOTE: Read in capacity data, and existing generation for solar and wind. Read in node data for connecting plants to grid
        for gen_type in ["solar", "wind"]:
            capacity_df = pd.read_csv(output_path / f"capacity_{gen_type}.csv", header=None)
            capacity_recon_df = pd.read_csv(output_path / f"capacity_{gen_type}_recon.csv", header=None)
            extant_df = pd.read_csv(files_needed_for_tool / f"extant_{gen_type}.csv", header=None)
            nodes = node_formatter(province)
            # NOTE: capacity_count_df is the same as capacity_df
            capacity_count_df = pd.read_csv(output_path / f"capacity_{gen_type}.csv", header=None)
            
            prev_period = periods[0]
            # Sum old (capacity_df) and new (capacity_recon) capacity
            for index,row in capacity_df.iterrows():
                capacity_df.at[index, 2] = capacity_df.at[index, 2] + capacity_recon_df.at[index, 2]
                capacity_count_df.at[index, 2] = capacity_df.at[index, 2] + capacity_recon_df.at[index, 2] + extant_df.at[index, 2]
                
            for index,row in capacity_df.loc[capacity_count_df[2] != 0].iterrows():
                
                period_f = row[0].replace('(','').replace("'", '') # NOTE: What does period_f mean?
                
                if prev_period != period_f and gen_type == "solar":
                    period_index = periods.index(prev_period)
                    vre_model_input_df = df_from_excel(model_inputs_out / f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                elif prev_period != period_f and gen_type == "wind":
                    period_index = periods.index(period_f)
                    vre_model_input_df_solar = df_from_excel(model_inputs_out / f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                    period_index = periods.index(prev_period)
                    vre_model_input_df = df_from_excel(model_inputs_out / f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                    vre_model_input_df = pd.concat([vre_model_input_df_solar, vre_model_input_df], ignore_index=True)
                else:
                    period_index = periods.index(period_f)
                    vre_model_input_df = df_from_excel(model_inputs_out / f"model inputs - {province}_{periods[period_index]}.xlsx", "vre plants")
                
                vre_model_input_df = vre_model_input_df.drop_duplicates(subset=['name'])
                grid_cell = row[1].replace(')','').replace("'", '')

                # Conditionally enter this loop if the province is found in the gl_to_ba_df dataframe
                if(gl_to_ba_df.at[index%2278, 1].split(".")[0] == provinces_full[province]):
                    
                    capacity = extant_df.at[index, 2] + capacity_df.iat[index, 2]
                    
                    latitude = None
                    longitude = None
                    balancing_area = None
                    
                    # TODO: Easy to remove this loop
                    for j,row in coordinate_df.loc[coordinate_df['grid cell'] == int(grid_cell)].iterrows():
                        latitude = row['lat']
                        longitude = row['lon']
                
                    # Needed for national model
                    for j,row in gl_to_ba_df.loc[gl_to_ba_df[0] == int(grid_cell)].iterrows():
                        balancing_area = row[1]

                    national_bus = f"{province}.{balancing_area.split('.')[1]}"
                    
                    # Initial location setting for distance function
                    coords_1 = (latitude, longitude)
                    coords_2 = (nodes.at[0, "Latitude"], nodes.at[0, "Longitude"])
                    min = geodesic(coords_1, coords_2).km
                    i = 0
                    
                    # TODO: Could probably replace this with my distance finding function
                    for j,row in nodes.iterrows(): # Connecting vre plants to nearest bus
                        if(str(row["Latitude"]) != "nan"):
                            coords_2 = (row["Latitude"], row["Longitude"])
                            temp = geodesic(coords_1, coords_2).km
                            if(min > temp):
                                min = temp
                                i = j
                    
                    bus = nodes.at[i, "bus"]

                    temp = vre_model_input_df.loc[vre_model_input_df["latitude"] == latitude]
                    temp = temp.loc[temp['kind'] == gen_type]
                    
                    if len(temp.loc[temp['longitude'] == longitude]) == 0:
                        
                        suffix = 1
                        name = f"{gen_type[0].upper()}{gen_type[1:]}_{suffix}" # What?
                        
                        while len(vre_model_input_df.loc[vre_model_input_df['name'] == name]) != 0:
                            suffix += 1
                            name = f"{gen_type[0].upper()}{gen_type[1:]}_{suffix}" 
                        merra_lat = round(2*(latitude+90),0)
                        merra_long = round(1.5*(longitude+180),0)                        
                        values = [gen_type.capitalize(), name, gen_type, latitude, longitude, capacity, bus, bus, None, None, merra_lat, merra_long]
                        
                        new_row = pd.DataFrame([values], columns=vre_model_input_df.columns)
                        vre_model_input_df = pd.concat([vre_model_input_df, new_row], ignore_index=True)
                        del temp
                    else:

                        # temp = vre_model_input_df.loc[vre_model_input_df["latitude"] == latitude]
                        temp = temp.loc[temp['longitude'] == longitude]

                        # for j,row in temp.iterrows():
                            # temp = temp.loc[temp['longitude'] == longitude]
                            
                        for g,row in temp.iterrows():
                            
                            vre_model_input_df.at[g, "[MW]"] += capacity
                            if vre_model_input_df.at[g, "[MW]"] < 0:
                                vre_model_input_df.at[g, "[MW]"] = 0
                        del temp
            
                wb_target = load_workbook(model_inputs_out / f"model inputs - {province}_{period_f}.xlsx", data_only=True)
                #wb_target = load_workbook(f"model inputs - {province}_{periods[i-1]}.xlsx", data_only=True)
                del wb_target["vre plants"]
                writer = pd.ExcelWriter(model_inputs_out / f"model inputs - {province}_{period_f}.xlsx", engine='openpyxl')
                #writer = pd.ExcelWriter(f"model inputs - {province}_{periods[i-1]}.xlsx", engine='openpyxl')
                writer.book = wb_target

                vre_model_input_df.to_excel(writer, sheet_name="vre plants", index=False)
                writer.save()
                wb_target.close() 
                del vre_model_input_df
                prev_period = period_f

# TODO: Change path of input VRE files to location which will work on other machines
def more_vre():

    for period in periods:
        for province in provinces:

            df = pd.read_excel(model_inputs_out / f"model inputs - {province}_{period}.xlsx", sheet_name="vre plants", usecols= ['kind', 'latitude_MERRA', 'longitude_MERRA'])
            df.columns = ['kind', 'lat', 'long']

            # Copies needed VRE analysis data from master folder to province/period specific folder
            for index, row in df.iterrows():

                dest = f"results\\vre\\{province}_{period}\\{row.kind.capitalize()}_Generation_Data\\{row['lat']}-{row['long']}"

                if not os.path.isdir(dest):

                    os.makedirs(dest)

                shutil.copyfile(f"C:\\Users\\Noah\\Documents\\Summer 2022 Co-op\\VRE_Resource_Analysis\\{row.kind.capitalize()}_Generation_Data\\{row['lat']}-{row['long']}\\{row.kind.capitalize()}_Generation_Data_{row['lat']}-{row['long']}_2018.csv", dest + f"\\{row.kind.capitalize()}_Generation_Data_{row['lat']}-{row['long']}_2018.csv")
                shutil.copyfile(f"C:\\Users\\Noah\\Documents\\Summer 2022 Co-op\\VRE_Resource_Analysis\\{row.kind.capitalize()}_Generation_Data\\{row['lat']}-{row['long']}\\{row.kind.capitalize()}_Generation_Data_{row['lat']}-{row['long']}_2018.csv", dest + f"\\{row.kind.capitalize()}_Generation_Data_{row['lat']}-{row['long']}_2018.csv")

# NOTE: This function reads aggregated non-vre data from COPPER (from results summary, skipping hydro which is handled separately), 
#       and for each period requested, either removes plants from the input model inputs sheet or distributes additional capacity amongst the existing plants.
#       The function does not add plants.
def add_non_VRE():

    result_summary_df = pd.read_excel(output_path / "Results_summary.xlsx", sheet_name="ABA_generation_mix")
    province_results = {}

    # Read in binaries for chosen new hydro development
    daily_binary = pd.read_csv(output_path / "day_renewal_binary.csv", 
                            names=['Year', 'Site', 'value'],
                            index_col=['Year'],
                            converters={'Year': lambda x: x.replace("('", '').replace("'", ""),
                                        'Site': lambda x: x.replace("')", "").replace(" '", "")})

    run_binary = pd.read_csv(output_path / "ror_renewal_binary.csv",
                            names=['Year', 'Site', 'value'],
                            index_col=['Year'],
                            converters={'Year': lambda x: x.replace("('", '').replace("'", ""),
                                        'Site': lambda x: x.replace("')", "").replace(" '", "")})

    month_binary = pd.read_csv(output_path / "month_renewal_binary.csv",
                            names=['Year', 'Site', 'value'],
                            index_col=['Year'],
                            converters={'Year': lambda x: x.replace("('", '').replace("'", ""),
                                        'Site': lambda x: x.replace("')", "").replace(" '", "")})

    chosen_hydro = pd.concat([daily_binary.loc[daily_binary['value'] == 1, 'Site'],
                            month_binary.loc[month_binary['value'] == 1, 'Site'],
                            run_binary.loc[run_binary['value']==1, 'Site']])

    
    for province in provinces:
        result_columns = []
        i = -1
        # Iterate over the columns of ABA_generation_mix
        for columnName, columnData in result_summary_df.iteritems():
            if columnName.split('.')[0] == provinces_full[province]:
                if columnName.split('.')[1] == "b":
                    result_columns[i] += columnData
                else:
                    # Add aggregated non-vre data from given province
                    result_columns.append(columnData)
                    i += 1
        province_results[province] = result_columns
               
    print("##### Adding to Non-VRE #####")
    for province in provinces:
        print(f'Calculating non-VRE for {province}')
        # Each provincial entry of province_results is the transposed columns of that provinces aggregate non-vre capacity (from COPPER)
        # There is one row for each of the years 2030, 2040 and 2050 (how can this data be used if we have different periods that we want to create SILVER inputs for?)
        for i, copper_non_vre in enumerate(province_results[province]):
            period = periods[i]
            # Reads from model inputs sheet which is created earlier (in add_vre_capacity)
            model_input_df = df_from_excel(model_inputs_out / f"model inputs - {province}_{period}.xlsx", "non-vre plants")

            # Remove hydro from model inputs
            model_input_df = model_input_df.loc[~model_input_df['kind'].str.contains(
                'hydro')]

            # copper_non_vre is aggregated non-vre capacities from COPPER, k is total for specific generation type
            for index, k in enumerate(copper_non_vre):
                result = k
                # NOTE: Still using the results summary excel file, shouldn't the indices have been transferred to the province_results dataframe?
                gen_type = result_summary_df.iat[index, 0]
                # Skip solar and wind
                if gen_type in ("solar", "wind"):
                    continue
                # Rename non-vre types for SILVER
                if gen_type == "gasCC":
                    gen_type = "NG_CT"
                if gen_type == "gasSC":
                    gen_type = "NG_CG"
                if gen_type == "gasccs":
                    gen_type = "NG_CC"
                # if gen_type == "biomass":
                #     gen_type = "biogas"

                # Subtract SILVER's total from COPPER's total (for given generation type)
                k -= model_input_df.loc[model_input_df['kind']
                                        == gen_type]["[MW]"].sum()
                
                # This algorithm works in the following way: 
                # If the COPPER capacity for a given generation type is greater than SILVER's in the base year, 
                # the tool adds ALL of the excess capacity onto the generator with the smallest capacity in the
                # base year for that type, and leaves the others the same (might be a good idea to change this 
                # so that it distributes the capacity instead). If the COPPER capacity is less than SILVER's, 
                # it deletes generators from SILVER in order of increasing capacity, until the capacity in SILVER
                # matches.
                while k != 0:
                    smallest_capacity = None
                    generator_to_remove = None
                    
                    # This loop looks for the generator (of kind == gen_type) with the smallest capacity
                    for j, generator in model_input_df.loc[model_input_df['kind'] == gen_type].loc[model_input_df["[MW]"] > 0].iterrows():
                        if smallest_capacity is None:
                            smallest_capacity = generator.get(5)
                            generator_to_remove = j
                        elif smallest_capacity > generator.get(5):
                            smallest_capacity = generator.get(5)
                            generator_to_remove = j
                        else:
                            continue

                    if smallest_capacity is None:
                        break
                    elif k < 0:
                        if abs(k) > smallest_capacity:
                            model_input_df.at[generator_to_remove, "[MW]"] = 0
                            k = k + smallest_capacity
                        else:
                            model_input_df.at[generator_to_remove,
                                              "[MW]"] = smallest_capacity + float(k)
                            k = 0
                    elif k >= 0:
                        model_input_df.at[generator_to_remove,
                                          "[MW]"] += float(k)
                        k = 0
                    
            model_input_df = model_input_df.loc[model_input_df["[MW]"] > 0]

            ## Hydro
            data_inventory = pd.read_excel(coders_path / f'210516-{province}-DataInventory.xlsx', sheet_name='Generation', header = 1).dropna(how='all')
            
            # Filter to only include hydro plants surviving to given year range
            data_inventory = data_inventory.loc[data_inventory['End Year'] >= int(period)+1]

            # New hydro builds
            new_hydro = pd.read_excel(files_needed_for_tool / "hydro_new_recon_nopump_with_coords.xlsx").dropna()
            new_hydro = new_hydro.loc[new_hydro['Balancing Area'].str.contains(provinces_full[province])]

            if int(period) in chosen_hydro.index:

                new_hydro = new_hydro.loc[new_hydro['Short Name'].isin(chosen_hydro[int(period)])]

                # Formatting to match old_plants
                new_hydro = new_hydro[['Project Name', 'Type', 'Latitude', 'Longitude', 'Additional Capacity (MW)']]
                new_hydro.columns = ['Project Name', 'Generation Type - COPPER', 'Latitude', 'Longitude', 'Installed Capacity']
            
            else:

                # Clear the dataframe if there is no new hydro for this period
                new_hydro.drop(new_hydro.index, inplace=True)


            if data_inventory['Generation Type - COPPER'].str.contains('hydro').any():

                old_plants = plants_filter(data_inventory)

                if not new_hydro.empty:

                    plants = pd.concat([old_plants, new_hydro], ignore_index=True)

                else:

                    plants = old_plants.copy()

                nodes = node_formatter(province)

                # Rename hydro run to daily
                if plants['Generation Type - COPPER'].str.contains('hydro_run').any():

                    plants.loc[plants['Generation Type - COPPER'] == 'hydro_run', 'Generation Type - COPPER'] = 'hydro_hourly'
                
                # Loop to find closest connections for plants
                plants['Connected Bus'] = np.nan

                # NOTE:Might be able to used df.apply instead
                for index, plant in plants.iterrows():

                    min, min_index = min_finder(plant, nodes)

                    plants.loc[index, 'Connected Bus'] = nodes.loc[min_index, 'bus']

                plants = plants.set_index('Generation Type - COPPER').sort_index().reset_index()

                plants.columns = ['kind', 'Project Name', 'Latitude', 'Longitude', '[MW]', 'bus']

                # Numbering hydro facilities
                for hydro_type in plants['kind'].unique():

                    plants.loc[plants['kind'] == hydro_type, 'plant ID'] = [f"{hydro_type}_{x}" for x in range(1, len(plants.loc[plants['kind'] == hydro_type]) + 1)]

                # Adding some more data
                plants['name'] = plants['plant ID']
                plants['cost curve equation'] = '3.3P'
                plants['pmin'] = 0

                plants = plants[['kind', 'plant ID', 'name', '[MW]', 'bus', 'pmin', 'cost curve equation']]

                # Add data to model inputs
                model_input_df = pd.concat([model_input_df, plants])
                
                hydro_cf = hydro_capacities()[provinces_full[province]]
                plants.set_index('kind', inplace = True)

                hourly_list, daily_list, monthly_list = [], [], []
                
                for kind, row in plants.iterrows():

                    # Hydro run renamed to daily
                    if kind == 'hydro_hourly':

                        mw = row['[MW]']

                        hourly_cf = hydro_cf
                        # hourly_cf.index = hourly_cf.index.get_level_values('hour')

                        values = mw * hourly_cf
                        values.name = row['name']
                        values.index = pd.date_range(start = '1/1/2018', end = '12/31/2018 23:00', periods = 8760, name = 'date')

                        hourly_list.append(values)

                    # Hydro daily
                    if kind == 'hydro_daily':
                        
                        mw = row['[MW]']

                        daily_cf = hydro_cf.groupby(level=[0,1], sort=False).sum().reset_index(drop=True)
                        
                        values = mw * daily_cf
                        values.name = row['name']
                        values.index = [f"day_{x}" for x in range(1,366)]
                        values.index.name = 'date'
                
                        # hydro_daily.insert(len(hydro_daily.columns), row['name'], values)
                        daily_list.append(values)

                    # Hydro monthly
                    if kind == 'hydro_monthly':
                        
                        mw = row['[MW]']

                        monthly_cf = hydro_cf.groupby(level=[0], sort=False).sum().reset_index(drop=True)

                        values = mw * monthly_cf
                        values.name = row['name']
                        values.index = [f"month_{x}" for x in range(1, 13)]
                        values.index.name = 'date'

                        monthly_list.append(values)

                if len(hourly_list)>1:
                    hydro_hourly = pd.concat(hourly_list, axis=1)
                    hydro_hourly.to_csv(hydro_out / f"{period}_{province}_hydro_hourly.csv")

                if len(daily_list)>1:
                    
                    hydro_daily = pd.concat(daily_list, axis=1)
                    hydro_daily.to_csv(hydro_out / f"{period}_{province}_hydro_daily.csv")

                if len(monthly_list)>1:

                    hydro_monthly = pd.concat(monthly_list, axis=1)
                    hydro_monthly.to_csv(hydro_out / f"{period}_{province}_hydro_monthly.csv")


            else:
                print(f'No hydro for {province}_{period}, skipping ')

            ### Write the updates
            wb_target = load_workbook(model_inputs_out / f"model inputs - {province}_{period}.xlsx", data_only=True)
            del wb_target["non-vre plants"]
            writer = pd.ExcelWriter(model_inputs_out / f"model inputs - {province}_{period}.xlsx", engine='openpyxl')
            writer.book = wb_target
            model_input_df.to_excel(writer, sheet_name="non-vre plants", index=False)
            writer.save()

def add_storage(province):
    print("##### Adding storage ######")
    capacity_storage_df = pd.read_csv(output_path / "capacity_storage.csv", header=None)
    capacity_storage_df = capacity_storage_df.sort_values(by=[0])

    for i,row in capacity_storage_df.loc[capacity_storage_df[3] != 0].iterrows(): 
    
        period = str(row.get(0))
        prev_period = periods.index(period) - 1
        balancing_area = row.get(2)
        province_storage = balancing_area.split('.')[0]
        new_capacity = row.get(3)
        if row.get(1) == "LB":
            storage_type = "battery"
        else:
            storage_type = row.get(1)
        name = storage_type + "." + balancing_area.split('.')[1]
        if provinces_full[province] != province_storage:
            continue
        model_input_df = df_from_excel(model_inputs_out / f"model inputs - {province}_{periods[prev_period]}.xlsx", "storage")
        
        smallest_capacity = None
        effected_storage = None

        for j,storage in model_input_df.iterrows():
            if storage_type not in storage.get(0):
                continue
            if smallest_capacity is None:
                smallest_capacity = storage.get(4)
                effected_storage = j
            else:
                if smallest_capacity > storage.get(4):
                    effected_storage = j
                    smallest_capacity = storage.get(4)
        
        if effected_storage is None:
            kind = storage_type
            initial = 1
            storage_type = f"{storage_type}_{balancing_area}"
            while len(model_input_df.loc[model_input_df["plant ID"] == type] != 0):
                initial +=1
                storage_type = f"{storage_type}_{balancing_area}"
            storage_capacity_max = 0
            if(kind == "PHS"):
                storage_capacity_max = new_capacity*8
            elif(kind == "LB"):
                storage_capacity_max = new_capacity*4
                
            values = [name, name, kind, None, new_capacity, storage_capacity_max, 1, None, None, "0P", None, None, None, 0]
            
            new_row = pd.DataFrame([values], columns=model_input_df.columns.tolist())
            model_input_df = pd.concat([model_input_df, new_row], ignore_index=True)
        
            
        else:
            model_input_df.at[effected_storage, "[MW]"] = model_input_df.at[effected_storage, "[MW]"] +new_capacity
            if(kind == "PHS"):
                model_input_df.at[effected_storage, "storagecapacitymax"] = model_input_df.at[effected_storage, "[MW]"]*8
            elif(kind == "LB"):
                model_input_df.at[effected_storage, "storagecapacitymax"] = model_input_df.at[effected_storage, "[MW]"]*4
            
        wb_target = load_workbook(model_inputs_out / f"model inputs - {province}_{period}.xlsx", data_only=True)
        del wb_target["storage"]
        writer = pd.ExcelWriter(model_inputs_out / f"model inputs - {province}_{period}.xlsx", engine='openpyxl')
        writer.book = wb_target
        model_input_df.to_excel(writer, sheet_name="storage", index=False)
        writer.save()
        
def main():

    TESTING = False

    if TESTING:

        add_non_VRE()

    else:

        start_time = process_time()
        
        add_vre_capacity()
        vre_time = process_time()
        print(f"VRE took {vre_time-start_time} seconds")
        more_vre()

        add_non_VRE()
        non_vre_time = process_time()
        print(f"Non-VRE took {non_vre_time-vre_time} seconds")

        # TODO: province loop should be in the function like the others
        for province in provinces:
            add_storage(province)

        end_time = process_time()
        print("Total Time: --- %s seconds ---" % (end_time-start_time))
    
if __name__ == "__main__":
    main()