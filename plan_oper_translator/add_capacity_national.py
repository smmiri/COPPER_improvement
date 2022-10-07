from collections import defaultdict
from openpyxl.workbook import workbook
import pandas as pd
import numpy as np
import shutil
import math
import sys
import time
from openpyxl import load_workbook
#import xlwings as xw
import os

#### Global Variables #####
periods_list = ["2025", "2030", "2035", "2040", "2045", "2050"]

capacity_sum = {'2025': None, '2030':None, '2035':None, '2040':None, '2045':None, '2050':None}


PREVIOUS_YEAR = 2278 # each year is composed of 2278 inputs therefore we can use this number to get the previous years capacity for example index = 9678 -> (2045, 1785), index = 9678 - 2278 -> (2040, 1785)

provinces_full = {
            "British Columbia": "BC",
            "Alberta": "AB",
            "Manitoba": "MN",
            "New Brunswick": "NB",
            "Newfoundland and Labrador": "NL",
            "Nova Scotia": "NS",
            "Ontario": "ON",
            "Quebec": "QB",
            "Saskatchewan": "SK",
            "Prince Edward Island": "PE"
            }


provinces_full_1 = {
            "BC.a": "British Columbia.a",
            "AB.a": "Alberta.a",
            "MN.a": "Manitoba.a",
            "NB.a": "New Brunswick.a",
            "NL.a": "Newfoundland and Labrador.a" ,
            "NS.a": "Nova Scotia.a",
            "ON.a": "Ontario.a",
            "QB.a": "Quebec.a",
            "SK.a": "Saskatchewan.a",
            "PE.a": "Prince Edward Island.a", 
            "BC.b": "British Columbia.b",
            "AB.b": "Alberta.b",
            "MN.b": "Manitoba.b",
            "NB.b": "New Brunswick.b",
            "NL.b": "Newfoundland and Labrador.b" ,
            "NS.b": "Nova Scotia.b",
            "ON.b": "Ontario.b",
            "QB.b": "Quebec.b",
            "SK.b": "Saskatchewan.b",
            "PE.b": "Prince Edward Island.b"
            }


provinces_full_inverse = {
            "British Columbia.a": "BC.a",
            "Alberta.a": "AB.a",
            "Manitoba.a": "MN.a",
            "New Brunswick.a": "NB.a",
            "Newfoundland and Labrador.a": "NL.a",
            "Nova Scotia.a": "NS.a",
            "Ontario.a": "ON.a",
            "Quebec.a": "QB.a",
            "Saskatchewan.a": "SK.a",
            "Prince Edward Island.a": "PE.a",
            "British Columbia.b": "BC.b",
            "Alberta.b": "AB.b",
            "Manitoba.b": "MN.b",
            "New Brunswick.b": "NB.b",
            "Newfoundland and Labrador.b": "NL.b",
            "Nova Scotia.b": "NS.b",
            "Ontario.b": "ON.b",
            "Quebec.b": "QB.b",
            "Saskatchewan.b": "SK.b",
            "Prince Edward Island.b": "PE.b"
            }

periods = {
            0: "2025",
            1: "2030",
            2: "2035",
            3: "2040",
            4: "2045",
            5: "2050",
            }
"""
    This function ensures that any forumulas from the Excel sheet are read
    as values by Pandas and not as an empty cell.
"""
def df_from_excel(path, sheet_name):
    return pd.read_excel(path, header=0, sheet_name=sheet_name)
def make_df(input):
    results_summary_df = pd.read_excel(input, header=0, sheet_name="ABA_generation_mix", index_col=0).dropna()
    columns = ["BA","Type"]
    new_df = pd.DataFrame([], columns=columns)
    period_2025,period_2030,period_2035,period_2040,period_2045,period_2050  = [],[],[],[],[],[]
    for index,row in results_summary_df.iterrows():
        gen_type = index
        if gen_type == "waste":
            gen_type = "biomass"
        for column in row.index:
            period = column[-1]
            if period not in ('1','2','3','4','5'):
                period = periods[0]
                balancing_area = column
            else:
                period = periods[int(period)]
                balancing_area = column[:-2]
            
            capacity = row[column]
            if period == "2025":
                period_2025.append(capacity)
            elif period == "2030":
                period_2030.append(capacity)
            elif period == "2035":
                period_2035.append(capacity)
            elif period == "2040":
                period_2040.append(capacity)
            elif period == "2045":
                period_2045.append(capacity)
            elif period == "2050":
                period_2050.append(capacity)

            if len(new_df.loc[new_df['BA'] == balancing_area].loc[new_df['Type'] == gen_type]) == 0:
                values = (balancing_area, gen_type)
                new_row = pd.DataFrame([values], columns=columns)
                new_df = pd.concat([new_df,new_row], ignore_index=True)
    
    new_df["2025"] = period_2025
    new_df["2030"] = period_2030
    new_df["2035"] = period_2035
    new_df["2040"] = period_2040
    new_df["2045"] = period_2045
    new_df["2050"] = period_2050

    return new_df



def add_vre_capacity():
    print("##### Adding vre capacity ######")

    for period in periods_list:
        mock = pd.read_excel("model inputs - CA.xlsx", sheet_name="vre plants")
        vre_model_input_df = pd.DataFrame([], columns =mock.columns)
        wb_target = load_workbook(f"model inputs - CA_{period}.xlsx", data_only=True)
        del wb_target["vre plants"]
        writer = pd.ExcelWriter(f"model inputs - CA_{period}.xlsx", engine='openpyxl')
        writer.book = wb_target
        vre_model_input_df.to_excel(writer, sheet_name="vre plants", index=False)
        writer.save()
        wb_target.close()
    for gen_type in ["solar", "wind"]:
        capacity_df = pd.read_csv(f"capacity_{gen_type}.csv", header=None)
        capacity_recon_df = pd.read_csv(f"capacity_{gen_type}_recon.csv", header=None)
        extant_df = pd.read_csv(f"extant_{gen_type}.csv", header=None)
        coordinate_df = pd.read_excel("coordinate.xlsx", sheet_name="coordinate_system")
        gl_to_ba_df = pd.read_csv("map_gl_to_ba.csv", header=None)
        count = 0
        prev_period = "2025"
        
        for index,row in capacity_df.iterrows():
            '''
            Add capacity, capacity recon, and extant and we need to iterate and apply
            the formula (e.g SILVER_capacity(x,2035) = capacity(x,2025) + capacity(x,2030) + capacity(x,2035) + extant_capacity(x,2035))
            
            '''
            current_period = row[0].replace('(','').replace("'", '')

            if (current_period == '2025'):
                capacity = capacity_df.at[index, 2] + capacity_recon_df.at[index,2] + extant_df.at[index,2]
            elif (current_period == '2030'):
                capacity = capacity_df.at[index, 2] + capacity_recon_df.at[index,2] + extant_df.at[index,2] + capacity_df.at[index-PREVIOUS_YEAR, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR, 2]
            elif(current_period == '2035'):
                capacity = capacity_df.at[index, 2] + capacity_recon_df.at[index,2] + extant_df.at[index,2] + capacity_df.at[index-PREVIOUS_YEAR, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR, 2] + capacity_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*2, 2] 
            elif(current_period == '2040'):
                capacity = capacity_df.at[index, 2] + capacity_recon_df.at[index,2] + extant_df.at[index,2] + capacity_df.at[index-PREVIOUS_YEAR, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR, 2] + capacity_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_df.at[index-PREVIOUS_YEAR*3, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*3, 2]
            elif(current_period == '2045'):
                capacity = capacity_df.at[index, 2] + capacity_recon_df.at[index,2] + extant_df.at[index,2] + capacity_df.at[index-PREVIOUS_YEAR, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR, 2] + capacity_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_df.at[index-PREVIOUS_YEAR*3, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*3, 2] + capacity_df.at[index-PREVIOUS_YEAR*4, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*4, 2] 
            elif(current_period == '2050'):
                capacity = capacity_df.at[index, 2] + capacity_recon_df.at[index,2] + extant_df.at[index,2] + capacity_df.at[index-PREVIOUS_YEAR, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR, 2] + capacity_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*2, 2] + capacity_df.at[index-PREVIOUS_YEAR*3, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*3, 2]+ capacity_df.at[index-PREVIOUS_YEAR*4, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*4, 2] + capacity_df.at[index-PREVIOUS_YEAR*5, 2] + capacity_recon_df.at[index-PREVIOUS_YEAR*5, 2]
                

            if (capacity == 0):
                """
                do not want sums that equal zero outputted to the excel sheets
                """
                continue
            
            vre_model_input_df = df_from_excel(f"model inputs - CA_{current_period}.xlsx", "vre plants") #get excel file for current period
            vre_model_input_df = vre_model_input_df.drop_duplicates(subset=['name'])
            grid_cell = row[1].replace(')','').replace("'", '') #access grid cell from capacity_df (e.g '1182', '2277')

            for j,row in coordinate_df.loc[coordinate_df['grid cell'] == int(grid_cell)].iterrows():
                latitude = row['lat']
                longitude = row['lon']


            for j,row in gl_to_ba_df.loc[gl_to_ba_df[0] == int(grid_cell)].iterrows():
                balancing_area = row[1] #gets the bus of the current row(e.g 'British Columbia.a', 'Alberta.a')


                        
            bus = f"{provinces_full[balancing_area.split('.')[0]]}.{balancing_area.split('.')[1]}" #converts balancing_area to a abbreviation(e.g 'British Columbia.a' -> 'BC.a)

            suffix = 1
            name = f"{gen_type[0].upper()}{gen_type[1:]}_{suffix}"
            pi = f"{gen_type[0].upper()}{gen_type[1:]}"
            while len(vre_model_input_df.loc[vre_model_input_df['name'] == name]) != 0: #iterates until a the name with the suffix at the end does not exist(e.g Solar_1)
                suffix += 1
                name = f"{gen_type[0].upper()}{gen_type[1:]}_{suffix}" 
            merra_lat = round(2*(latitude+90),0)
            merra_long = round(1.5*(longitude+180),0)
            values = [pi, name, gen_type, latitude, longitude, capacity, bus, None, None, merra_lat, merra_long] #creates list with all the data for current row
            new_row = pd.DataFrame([values], columns=vre_model_input_df.columns)#creates a dataframe from the list containg all the data for the curent row
            vre_model_input_df = pd.concat([vre_model_input_df, new_row], ignore_index=True) #adds the new_row dataframe to the vre_model as a new a row





            '''
            Updates the sheet "vre plants" for the current period excel file with the current row
            '''

            wb_target = load_workbook(f"model inputs - CA_{current_period}.xlsx", data_only=True) 
            del wb_target["vre plants"]
            writer = pd.ExcelWriter(f"model inputs - CA_{current_period}.xlsx", engine='openpyxl')
            writer.book = wb_target
            vre_model_input_df.to_excel(writer, sheet_name="vre plants", index=False)
            writer.save()
            wb_target.close()
            del vre_model_input_df



        
            
def add_non_vre_capacity(capacity_df):
    print("##### Adding non vre capacity ######")
    for i,period in enumerate(periods):
        period = periods[period]
        
        if period == "2025":
            shutil.copy(f"model inputs - CA.xlsx", f"model inputs - CA_{period}.xlsx")
        else:
            shutil.copy(f"model inputs - CA_{periods[i-1]}.xlsx", f"model inputs - CA_{period}.xlsx")

        non_vre_model_input_df = df_from_excel(f"model inputs - CA_{period}.xlsx", "non-vre plants")
        non_vre_model_input_df = non_vre_model_input_df[0:0]

        for index,row in capacity_df.loc[capacity_df[period] != 0].iterrows():
            if row["Type"] in ("wind", "solar"):
                continue

            else:

                kind = row["Type"]
                name = f"{kind}_{provinces_full[row['BA'].split('.')[0]]}.{row['BA'][-1]}"
                bus = f"{provinces_full[row['BA'].split('.')[0]]}.{row['BA'][-1]}"
                capacity = row[period]
                if capacity < 0:
                    continue
                values = [name, name, kind, None, None, capacity, bus]
                for j in range(0,12):
                    values.append(None)
                new_row = pd.DataFrame([values], columns=non_vre_model_input_df.columns.tolist())
                non_vre_model_input_df = pd.concat([non_vre_model_input_df,new_row], ignore_index=True)
        
        wb_target = load_workbook(f"model inputs - CA_{period}.xlsx", data_only=True)
        del wb_target["non-vre plants"]
        writer = pd.ExcelWriter(f"model inputs - CA_{period}.xlsx", engine='openpyxl')
        writer.book = wb_target
        non_vre_model_input_df.to_excel(writer, sheet_name="non-vre plants", index=False)
        writer.save()  




def add_storage():
    print("##### Adding storage ######")
    capacity_storage_df = pd.read_csv("capacity_storage.csv", header=None)
    capacity_storage_df = capacity_storage_df.sort_values(by=[0])
    periods = ["2025", "2030", "2035", "2040", "2045", "2050"]
    prev_period = None
    for i,row in capacity_storage_df.loc[capacity_storage_df[3] != 0].iterrows(): 
    
        period = str(row.get(0))
    
        balancing_area = row.get(2)
        new_capacity = row.get(3)
        storage_type = row.get(1)

        if storage_type == 'PHS':
            kind = storage_type
        else:
            kind = 'batteries'

        
        if prev_period is None:
            model_input_df = df_from_excel(f"model inputs - CA_{period}.xlsx", "storage")
        elif period != periods[prev_period]:
            model_input_df = df_from_excel(f"model inputs - CA_{periods[prev_period]}.xlsx", "storage")
        else:
            model_input_df = df_from_excel(f"model inputs - CA_{period}.xlsx", "storage")

        storage_type = f"{storage_type}_{provinces_full[balancing_area.split('.')[0]]}.{balancing_area.split('.')[1]}"

        if len(model_input_df.loc[model_input_df["plant ID"] == storage_type]) == 0:
                        
            values = [storage_type, storage_type, kind, None, new_capacity, None, 1, None, f"{provinces_full[balancing_area.split('.')[0]]}.{balancing_area.split('.')[1]}"]
            for col in range(0,5):
                values.append(None)

            new_row = pd.DataFrame([values], columns=model_input_df.columns.tolist())
            model_input_df = pd.concat([model_input_df, new_row], ignore_index=True)
        
            
        else:
            effected_storage = None
            for x,storage in model_input_df.loc[model_input_df["plant ID"] == storage_type].iterrows():
                effected_storage = x
                
            model_input_df.at[effected_storage, "[MW]"] = model_input_df.at[effected_storage, "[MW]"] +new_capacity
        
        prev_period = periods.index(period) 
        wb_target = load_workbook(f"model inputs - CA_{period}.xlsx", data_only=True)
        del wb_target["storage"]
        writer = pd.ExcelWriter(f"model inputs - CA_{period}.xlsx", engine='openpyxl')
        writer.book = wb_target
        model_input_df.to_excel(writer, sheet_name="storage", index=False)
        writer.save()




def transmission_capacity():
    print("##### Adding transmission_capacity ######")
    for period in periods_list: #iterate throughout the periods
        capacity_transmission = pd.read_csv("capacity_transmission.csv", header=None) #gets the data that needs to be calculated
        # capacity_transmission.to_csv("/Users/armaanchhina/Desktop/tmp_conversion/capacity.csv")
        mock = pd.read_excel("model inputs - CA.xlsx", sheet_name="existing transmission") #creates a dataframe from model_inputs-CA to be able to use those columns in a new tmp dataframe
        existing_transmission_model_input_df = pd.DataFrame([], columns =mock.columns) #creates new dataframe mocking the columns from the mock dataframe
        existing_transmission = pd.read_excel(f"model inputs - CA_{period}.xlsx", header = 0, usecols='A:G', sheet_name="existing transmission") #the data that the formula will be applied to
        # existing_transmission.to_excel("/Users/armaanchhina/Desktop/tmp_conversion/vre.xlsx")
        # existing_transmission_model_input_df.to_excel("/Users/armaanchhina/Desktop/tmp_conversion/exist.xlsx")

        for index, row in existing_transmission.iterrows(): 
            name = row["name"]
            from_bus = row["from bus"]
            to_bus = row["to bus"]
            voltage = row["Voltage"]
            length = row["length"]
            reactance = row["reactance"]
            bus_from = f"{provinces_full_1[from_bus]}"
            bus_to = f"{provinces_full_1[to_bus]}"
            pmax = 0
            # pmax_base = row["pmax-base"]
            year = 0

            if period == "2025": #excutes the formula for 2025
                for index,row in capacity_transmission.iterrows():
                    year = row[0]
                    from_cap = row[1]
                    to_cap = row[2]
                    if(year>=2030):
                        # print("break")
                        break
                    if((from_cap == bus_from and to_cap == bus_to) or (from_cap == bus_to and to_cap == bus_from)):
                        # print("adding")
                        # print("index:", index)
                        pmax += row[3]
                        # print(pmax) 
                pmax = extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period)   

            if period == "2030": #excutes the formula for 2030
                for index,row in capacity_transmission.iterrows():
                    year = row[0]
                    from_cap = row[1]
                    to_cap = row[2]
                    if(year>=2035):
                        break
                    if((from_cap == bus_from and to_cap == bus_to) or (from_cap == bus_to and to_cap == bus_from)):
                        pmax += row[3]
                pmax = extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period)   


            if period == "2035": #excutes the formula for 2035
                for index,row in capacity_transmission.iterrows():
                    year = row[0]
                    from_cap = row[1]
                    to_cap = row[2]
                    if(year>=2040):
                        # print("break")
                        break
                    if((from_cap == bus_from and to_cap == bus_to) or (from_cap == bus_to and to_cap == bus_from)):
                        # print("adding")
                        # print("index:", index)
                        pmax += row[3]
                        # print(pmax)
                pmax = extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period)   

            if period == "2040": #excutes the formula for 2040
                for index,row in capacity_transmission.iterrows():
                    year = row[0]
                    from_cap = row[1]
                    to_cap = row[2]
                    if(year>=2045):
                        # print("break")
                        break
                    if((from_cap == bus_from and to_cap == bus_to) or (from_cap == bus_to and to_cap == bus_from)):
                        # print("adding")
                        # print("index:", index)
                        pmax += row[3]
                        # print(pmax)
                pmax = extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period)   

            if period == "2045": #excutes the formula for 2045
                for index,row in capacity_transmission.iterrows():
                    year = row[0]
                    from_cap = row[1]
                    to_cap = row[2]
                    if(year>=2050):
                        # print("break")
                        break
                    if((from_cap == bus_from and to_cap == bus_to) or (from_cap == bus_to and to_cap == bus_from)):
                        # print("adding")
                        # print("index:", index)
                        pmax += row[3]
                        # print(pmax)
                pmax = extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period)   

            if period == "2050": #excutes the formula for 2050
                for index,row in capacity_transmission.iterrows():
                    year = row[0]
                    from_cap = row[1]
                    to_cap = row[2]
                    if(year>=2055):
                        # print("break")
                        break
                    if((from_cap == bus_from and to_cap == bus_to) or (from_cap == bus_to and to_cap == bus_from)):
                        # print("adding")
                        # print("index 2050:", index)
                        pmax += row[3]
                        # print(pmax)                
                pmax = extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period)   


            values = [name, from_bus, to_bus, voltage, length,pmax,reactance] #creates a list with all the data for the specific row we are iterating through
            new_row = pd.DataFrame([values], columns=existing_transmission_model_input_df.columns)#creates a dataframe from the list containg all the data for the curent row
            existing_transmission_model_input_df = pd.concat([existing_transmission_model_input_df, new_row], ignore_index=True)
            # if(period == '2030'):
            #      print('done')
            #      existing_transmission_model_input_df.to_excel("/Users/armaanchhina/Desktop/tmp_conversion/2030sum.xlsx")


            wb_target = load_workbook(f"model inputs - CA_{period}.xlsx", data_only=True) 
            del wb_target["existing transmission"]
            writer = pd.ExcelWriter(f"model inputs - CA_{period}.xlsx", engine='openpyxl')
            writer.book = wb_target
            existing_transmission_model_input_df.to_excel(writer, sheet_name="existing transmission", index=False)
            writer.save()
            wb_target.close()



def extant_value(pmax, year, from_cap,bus_from,to_cap,bus_to,period): #helper function for transmission_capacity
    extant_transmission = pd.read_csv("extant_transmission.csv")
    for index,row in extant_transmission.iterrows():
        name = row["ABA"]
        n = 2
        groups = name.split(".")
        value = (".".join(groups[:n]), ".".join(groups[n:]))
        nameFront = value[0]
        nameBack = value[1]
        if((bus_from == nameFront and bus_to == nameBack) or (bus_from == nameBack and bus_to == nameFront)):
            pmax += row[period]
    return pmax




def main():
    print('You need to have these files already available in the folfer:\n1-Coordinates.xls,\n2-map_gl_to_ba.csv,\n'
          '3-all extant files\n')
    new_path = input("Enter the directory's name containing input files:\n(i.e. CA_date_use)\n")
    os.chdir(os.getcwd() + "/" +  new_path)
    #filename = input("Enter the name of the COPPER output file:\n(i.e. Results_summary.xlsx)\n")
    start_time = time.time()
    formatted_df = make_df("Results_summary.xlsx")
    add_non_vre_capacity(formatted_df)

    add_storage()
    add_vre_capacity()

    transmission_capacity()
    print("--- %s seconds ---" % (time.time() - start_time))
    
if __name__ == "__main__":
    main()
